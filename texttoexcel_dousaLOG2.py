import re
import os
import openpyxl
from openpyxl.styles import Font
from datetime import datetime

input_directory = "D:\Dドキュメント\業務\D棟\参考資料\稼働実績ログ\M棟の実ログコピー\programtest2\R2"  # 入力ディレクトリを指定
output_directory = "D:\Dドキュメント\業務\D棟\参考資料\稼働実績ログ\M棟の実ログコピー\programtest2\R2\excelLog_kai4"  # 出力ディレクトリを指定

if not os.path.exists(output_directory):
    os.makedirs(output_directory)

for filename in os.listdir(input_directory):
    if filename.endswith(".log"):
        input_path = os.path.join(input_directory, filename)
        output_path = os.path.join(output_directory, f"{os.path.splitext(filename)[0]}.xlsx")

        # Excelファイルを作成
        wb = openpyxl.Workbook()
        ws = wb.active

        with open(input_path, 'r', encoding='utf-8') as file:
            for line in file:
                line = line.strip()
                print(f"Processing line: '{line}'")

                # 行A
                if re.match(r'^[-+]?\d+(\.\d+)?(\s+[-+]?\d+(\.\d+)?)*$', line):
                    print("Matched Row A")
                    values = list(map(float, line.split()))
                    ws.append(values)

                # 行B
                elif re.match(r'^[^\d\s]+$', line):
                    print("Matched Row B")
                    ws.append([line])
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)

                # 行C
                elif re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}\.\d{3}$', line):
                    print("Matched Row C")
                    dt = datetime.strptime(line, '%Y-%m-%d %H:%M:%S.%f')
                    row = [dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second, dt.microsecond // 1000]
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)

                # 行D
                elif re.match(r'^P\d{1,4}J\d{1,4}\(\d{1,4}\)S\d{1,4}T\d{1,4}N\d{1,4}M\d{1,4}$', line):
                    print("Matched Row D")
                    values = [line[0:2], line[2:7], line[7:9], line[9:11], line[11:13], line[13:16]]
                    ws.append(values)

                # 行E
                elif line.startswith('/'):
                    print("Matched Row E")
                    ws.append([line])

                # どの行にも一致しない場合
                else:
                    print("No matching row type")

        # Excelファイルを保存
        wb.save(output_path)
